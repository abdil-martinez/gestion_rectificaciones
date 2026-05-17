import io
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.db.models import Count, Avg, F, Q, ExpressionWrapper, DurationField
from django.utils import timezone
from django.http import HttpResponse


class DashboardView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.solicitudes.models import Solicitud, ESTADO_CHOICES
        from apps.accounts.models import CustomUser

        hoy = timezone.now().date()
        total = Solicitud.objects.count()

        por_estado = {}
        for codigo, label in ESTADO_CHOICES:
            por_estado[codigo] = {
                'label': label,
                'total': Solicitud.objects.filter(estado=codigo).count(),
            }

        vencidas = Solicitud.objects.filter(
            fecha_limite__lt=hoy
        ).exclude(estado__in=['FIN', 'ANU', 'APRO']).count()

        por_regional = list(
            Solicitud.objects.filter(regional__isnull=False)
            .values('regional__nombre')
            .annotate(total=Count('id'))
            .order_by('-total')[:10]
        )

        por_tipo_regional_estado = list(
            Solicitud.objects
            .values('regional__tipo_regional__nombre', 'estado')
            .annotate(total=Count('id'))
            .order_by('regional__tipo_regional__nombre', 'estado')
        )

        top_causales = list(
            Solicitud.objects.filter(tipo_causal__isnull=False)
            .values('tipo_causal__nombre')
            .annotate(total=Count('id'))
            .order_by('-total')[:5]
        )

        ultimas_10 = Solicitud.objects.select_related(
            'asegurado', 'tipo_causal', 'regional', 'analista_asignado'
        ).order_by('-created_at')[:10]

        from apps.solicitudes.serializers import SolicitudListSerializer
        ultimas_data = SolicitudListSerializer(ultimas_10, many=True, context={'request': request}).data

        tiempo_promedio = None
        solicitudes_finalizadas = Solicitud.objects.filter(
            estado='FIN',
            fecha_recepcion__isnull=False,
            fecha_resolucion__isnull=False,
        )
        if solicitudes_finalizadas.exists():
            from django.db.models.functions import ExtractDay
            tiempos = [
                (sol.fecha_resolucion - sol.fecha_recepcion).days
                for sol in solicitudes_finalizadas
                if sol.fecha_resolucion and sol.fecha_recepcion
            ]
            if tiempos:
                tiempo_promedio = sum(tiempos) / len(tiempos)

        return Response({
            'total_solicitudes':          total,
            'por_estado':                 por_estado,
            'solicitudes_vencidas':       vencidas,
            'tiempo_promedio_atencion':   tiempo_promedio,
            'solicitudes_por_regional':   por_regional,
            'top_causales':               top_causales,
            'ultimas_solicitudes':        ultimas_data,
            'por_tipo_regional_estado': [
                {
                    'tipo_regional': row['regional__tipo_regional__nombre'] or 'Sin tipo',
                    'estado':        row['estado'],
                    'total':         row['total'],
                }
                for row in por_tipo_regional_estado
            ],
        })


class ReporteProductividadView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.solicitudes.models import Solicitud
        from apps.accounts.models import CustomUser

        fecha_desde = request.query_params.get('fecha_desde')
        fecha_hasta = request.query_params.get('fecha_hasta')

        qs = Solicitud.objects.filter(analista_asignado__isnull=False)
        if fecha_desde:
            qs = qs.filter(created_at__date__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(created_at__date__lte=fecha_hasta)

        por_analista = list(
            qs.values(
                'analista_asignado__id',
                'analista_asignado__first_name',
                'analista_asignado__last_name',
                'analista_asignado__username',
            )
            .annotate(
                total=Count('id'),
                aprobadas=Count('id', filter=Q(estado='APRO')),
                rechazadas=Count('id', filter=Q(estado='RECH')),
                finalizadas=Count('id', filter=Q(estado='FIN')),
            )
            .order_by('-total')
        )

        result = []
        for item in por_analista:
            nombre = f"{item['analista_asignado__first_name']} {item['analista_asignado__last_name']}".strip()
            if not nombre:
                nombre = item['analista_asignado__username']
            result.append({
                'analista_id':  item['analista_asignado__id'],
                'analista':     nombre,
                'total':        item['total'],
                'aprobadas':    item['aprobadas'],
                'rechazadas':   item['rechazadas'],
                'finalizadas':  item['finalizadas'],
            })

        return Response({'productividad': result})


class ReporteCausalView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.solicitudes.models import Solicitud

        fecha_desde = request.query_params.get('fecha_desde')
        fecha_hasta = request.query_params.get('fecha_hasta')

        qs = Solicitud.objects.filter(tipo_causal__isnull=False)
        if fecha_desde:
            qs = qs.filter(created_at__date__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(created_at__date__lte=fecha_hasta)

        por_causal = list(
            qs.values('tipo_causal__id', 'tipo_causal__nombre', 'tipo_causal__tipo')
            .annotate(
                total=Count('id'),
                aprobadas=Count('id', filter=Q(estado='APRO')),
                rechazadas=Count('id', filter=Q(estado='RECH')),
                finalizadas=Count('id', filter=Q(estado='FIN')),
                monto_total=Count('monto_total'),
            )
            .order_by('-total')
        )

        return Response({'por_causal': por_causal})


class ReporteTipoRegionalView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.solicitudes.models import Solicitud

        fecha_desde = request.query_params.get('fecha_desde')
        fecha_hasta = request.query_params.get('fecha_hasta')

        qs = Solicitud.objects.filter(regional__isnull=False)
        if fecha_desde:
            qs = qs.filter(created_at__date__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(created_at__date__lte=fecha_hasta)

        por_tipo = list(
            qs.values(
                'regional__tipo_regional__id',
                'regional__tipo_regional__nombre',
            )
            .annotate(
                total=Count('id'),
                finalizadas=Count('id', filter=Q(estado='FIN')),
                aprobadas=Count('id',   filter=Q(estado='APRO')),
                rechazadas=Count('id',  filter=Q(estado='RECH')),
            )
            .order_by('-total')
        )

        por_regional = list(
            qs.values(
                'regional__tipo_regional__nombre',
                'regional__id',
                'regional__nombre',
            )
            .annotate(
                total=Count('id'),
                finalizadas=Count('id', filter=Q(estado='FIN')),
                aprobadas=Count('id',   filter=Q(estado='APRO')),
            )
            .order_by('regional__tipo_regional__nombre', '-total')
        )

        result_tipo = [
            {
                'id':         row['regional__tipo_regional__id'],
                'nombre':     row['regional__tipo_regional__nombre'] or 'Sin tipo',
                'total':      row['total'],
                'finalizadas': row['finalizadas'],
                'aprobadas':  row['aprobadas'],
                'rechazadas': row['rechazadas'],
            }
            for row in por_tipo
        ]

        result_regional = [
            {
                'tipo_regional': row['regional__tipo_regional__nombre'] or 'Sin tipo',
                'regional_id':   row['regional__id'],
                'regional':      row['regional__nombre'],
                'total':         row['total'],
                'finalizadas':   row['finalizadas'],
                'aprobadas':     row['aprobadas'],
            }
            for row in por_regional
        ]

        return Response({
            'por_tipo_regional': result_tipo,
            'por_regional':      result_regional,
        })


class ExportarExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from apps.solicitudes.models import Solicitud, ESTADO_CHOICES, PRIORIDAD_CHOICES

        tipo_reporte = request.query_params.get('tipo', 'solicitudes')
        fecha_desde  = request.query_params.get('fecha_desde')
        fecha_hasta  = request.query_params.get('fecha_hasta')

        wb = openpyxl.Workbook()
        ws = wb.active

        ORO_HEX  = 'CBab58'
        NAVY_HEX = '0F1932'

        header_fill  = PatternFill('solid', fgColor=ORO_HEX)
        header_font  = Font(bold=True, color=NAVY_HEX)
        header_align = Alignment(horizontal='center', vertical='center')
        thin_border  = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        if tipo_reporte == 'productividad':
            ws.title = 'Productividad'
            headers  = ['Analista', 'Total', 'Aprobadas', 'Rechazadas', 'Finalizadas']
            qs = Solicitud.objects.filter(analista_asignado__isnull=False)
            if fecha_desde:
                qs = qs.filter(created_at__date__gte=fecha_desde)
            if fecha_hasta:
                qs = qs.filter(created_at__date__lte=fecha_hasta)

            data = list(
                qs.values(
                    'analista_asignado__first_name',
                    'analista_asignado__last_name',
                    'analista_asignado__username',
                )
                .annotate(
                    total=Count('id'),
                    aprobadas=Count('id', filter=Q(estado='APRO')),
                    rechazadas=Count('id', filter=Q(estado='RECH')),
                    finalizadas=Count('id', filter=Q(estado='FIN')),
                )
                .order_by('-total')
            )

            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.fill = header_fill; cell.font = header_font
                cell.alignment = header_align; cell.border = thin_border

            for row_idx, item in enumerate(data, 2):
                nombre = f"{item['analista_asignado__first_name']} {item['analista_asignado__last_name']}".strip()
                if not nombre:
                    nombre = item['analista_asignado__username']
                row_data = [nombre, item['total'], item['aprobadas'], item['rechazadas'], item['finalizadas']]
                for col_idx, v in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=v)
                    cell.border = thin_border
        else:
            ws.title = 'Solicitudes'
            headers  = [
                'N° Solicitud', 'Estado', 'Prioridad', 'Asegurado', 'CI',
                'Empleador', 'Tipo Causal', 'Regional', 'Analista',
                'Fecha Recepción', 'Fecha Límite', 'Monto',
            ]
            qs = Solicitud.objects.select_related(
                'asegurado', 'empleador', 'tipo_causal', 'regional', 'analista_asignado'
            )
            if fecha_desde:
                qs = qs.filter(created_at__date__gte=fecha_desde)
            if fecha_hasta:
                qs = qs.filter(created_at__date__lte=fecha_hasta)

            estado_map   = dict(ESTADO_CHOICES)
            prioridad_map = dict(PRIORIDAD_CHOICES)

            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.fill = header_fill; cell.font = header_font
                cell.alignment = header_align; cell.border = thin_border

            for row_idx, sol in enumerate(qs, 2):
                row_data = [
                    sol.numero_solicitud,
                    estado_map.get(sol.estado, sol.estado),
                    prioridad_map.get(sol.prioridad, sol.prioridad),
                    sol.asegurado.nombre_completo if sol.asegurado else '',
                    sol.asegurado.cedula if sol.asegurado else '',
                    sol.empleador.nombre_razon_social if sol.empleador else '',
                    sol.tipo_causal.nombre if sol.tipo_causal else '',
                    sol.regional.nombre if sol.regional else '',
                    sol.analista_asignado.get_full_name() if sol.analista_asignado else '',
                    sol.fecha_recepcion,
                    sol.fecha_limite,
                    float(sol.monto_total),
                ]
                for col_idx, v in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=v)
                    cell.border = thin_border

        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"reporte_{tipo_reporte}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
