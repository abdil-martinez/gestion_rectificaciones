import io
from rest_framework import viewsets, status, filters, parsers
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
from django.utils import timezone

from .models import Solicitud, BitacoraSolicitud, Formulario, DocumentoRespaldo, Asegurado, Empleador, Solicitante
from .serializers import (
    SolicitudListSerializer, SolicitudDetailSerializer, SolicitudCreateSerializer,
    BitacoraSerializer, FormularioSerializer, DocumentoRespaldoSerializer,
    AseguradoSerializer, EmpleadorSerializer, SolicitanteSerializer,
)
from .workflow import puede_transitar, registrar_bitacora, transiciones_disponibles
from apps.catalogos.mixins import SoftDeleteMixin


def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


class SolicitudViewSet(SoftDeleteMixin, viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    filter_backends    = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields   = ['estado', 'regional', 'agencia', 'regional__tipo_regional', 'analista_asignado', 'prioridad', 'tipo_causal']
    search_fields      = ['numero_solicitud', 'asegurado__nombre', 'asegurado__cedula',
                          'asegurado__ap_paterno', 'asegurado__ap_materno']
    ordering_fields    = ['created_at', 'fecha_recepcion', 'fecha_limite', 'numero_solicitud', 'estado']
    ordering           = ['-created_at']

    def get_queryset(self):
        from django.db.models import Q
        qs = Solicitud.objects.select_related(
            'asegurado', 'empleador', 'solicitante',
            'tipo_solicitud', 'tipo_causal', 'regional', 'agencia',
            'analista_asignado', 'asignado_por', 'usuario_creador',
            'area_solicitante',
        )
        user = self.request.user
        fecha_desde  = self.request.query_params.get('fecha_desde')
        fecha_hasta  = self.request.query_params.get('fecha_hasta')
        mi_bandeja   = self.request.query_params.get('mi_bandeja') == 'true'

        if fecha_desde:
            qs = qs.filter(created_at__date__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(created_at__date__lte=fecha_hasta)

        exclude_estados = self.request.query_params.get('exclude_estados')
        if exclude_estados:
            qs = qs.exclude(estado__in=[e.strip() for e in exclude_estados.split(',')])

        # Filtro por plazo — umbrales dinámicos desde EstadoPlazo
        plazo = self.request.query_params.get('plazo')
        if plazo:
            import datetime
            from apps.catalogos.models import EstadoPlazo as EstadoPlazoModel
            today = timezone.now().date()
            estados_cerrados = ['FIN', 'ANU', 'RECT', 'RECH']

            if plazo == 'CERRADA':
                qs = qs.filter(estado__in=estados_cerrados)
            else:
                qs = qs.exclude(estado__in=estados_cerrados)
                # Umbrales ASC por limite_dias, sin "En Plazo" (es el total de días)
                umbrales = list(
                    EstadoPlazoModel.objects.exclude(nombre__iexact='En Plazo')
                    .order_by('limite_dias').values_list('nombre', 'limite_dias')
                )
                # Ej: [('Vencido',0),('Crítico',2),('Urgente',5),('Próximo a Vencer',10)]
                encontrado = False
                for i, (nombre, limite) in enumerate(umbrales):
                    if nombre.lower() == plazo.lower():
                        if i == 0:  # Vencido: fecha_limite anterior a hoy
                            qs = qs.filter(fecha_limite__lt=today)
                        else:
                            lower = umbrales[i - 1][1]
                            qs = qs.filter(
                                fecha_limite__gt=today + datetime.timedelta(days=lower),
                                fecha_limite__lte=today + datetime.timedelta(days=limite),
                            )
                        encontrado = True
                        break
                if not encontrado:
                    # "En Plazo": dias_restantes > último umbral
                    ultimo = umbrales[-1][1] if umbrales else 0
                    qs = qs.filter(fecha_limite__gt=today + datetime.timedelta(days=ultimo))

        todas            = self.request.query_params.get('todas') == 'true'
        analista_bandeja = self.request.query_params.get('analista_bandeja') == 'true'

        if mi_bandeja:
            qs = qs.filter(
                Q(usuario_creador=user, estado__in=['BOR', 'PEND', 'DEV', 'RECT', 'RECH'])
            )
        elif analista_bandeja:
            # Bandeja personal del analista:
            # - trabajo activo (asignado a él, en cualquier estado operativo)
            # - solicitudes procesadas/devueltas/anuladas donde él es el creador
            qs = qs.filter(
                Q(analista_asignado=user, estado__in=['ASIG', 'REV', 'DEV', 'RECT', 'RECH', 'ANU']) |
                Q(usuario_creador=user, estado__in=['DEV', 'RECT', 'RECH', 'ANU'])
            )
        elif user.rol == 'ANALIST' and not todas:
            if self.request.query_params.get('analista_asignado'):
                pass  # DjangoFilterBackend aplica analista_asignado=id directamente
            else:
                # El creador siempre puede ver sus propias solicitudes en cualquier estado
                qs = qs.filter(
                    Q(analista_asignado=user) |
                    Q(usuario_creador=user)
                )
        return qs

    def get_serializer_class(self):
        if self.action == 'create':
            return SolicitudCreateSerializer
        if self.action in ('list',):
            return SolicitudListSerializer
        return SolicitudDetailSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = request.user

        # Auto-asignar solicitante vinculado al usuario actual
        extra = {'usuario_creador': user}
        if not request.data.get('solicitante'):
            solicitante, _ = Solicitante.objects.get_or_create(
                usuario=user,
                defaults={
                    'nombre':     user.first_name or user.username,
                    'ap_paterno': user.last_name or '',
                }
            )
            extra['solicitante'] = solicitante

        solicitud = serializer.save(**extra)
        registrar_bitacora(
            solicitud=solicitud,
            usuario=user,
            estado_anterior=None,
            estado_nuevo='BOR',
            accion='CREACION',
            comentario='Solicitud creada.',
            ip=get_client_ip(request),
        )
        return Response(
            SolicitudDetailSerializer(solicitud, context={'request': request}).data,
            status=status.HTTP_201_CREATED,
        )

    def perform_create(self, serializer):
        pass

    @action(detail=True, methods=['post'], url_path='cambiar_estado')
    def cambiar_estado(self, request, pk=None):
        solicitud       = self.get_object()
        nuevo_estado    = request.data.get('estado')
        comentario      = request.data.get('comentario', '')
        analista_id     = request.data.get('analista_asignado')

        if not nuevo_estado:
            return Response({'detail': 'Debe indicar el nuevo estado.'}, status=status.HTTP_400_BAD_REQUEST)

        permitido, mensaje = puede_transitar(request.user.rol, solicitud.estado, nuevo_estado)
        if not permitido:
            return Response({'detail': mensaje}, status=status.HTTP_403_FORBIDDEN)

        estado_anterior    = solicitud.estado
        solicitud.estado   = nuevo_estado
        notif_estado_id    = request.data.get('notif_estado_id')
        notif_plantilla_id = request.data.get('notif_plantilla_id')
        if notif_estado_id is not None:
            solicitud.notif_estado_id    = notif_estado_id or None
            solicitud.notif_plantilla_id = notif_plantilla_id or None

        if analista_id and nuevo_estado == 'ASIG':
            from apps.accounts.models import CustomUser
            try:
                analista = CustomUser.objects.get(pk=analista_id)
                solicitud.analista_asignado = analista
            except CustomUser.DoesNotExist:
                pass

        if nuevo_estado == 'ASIG':
            solicitud.asignado_por = request.user

        if nuevo_estado == 'ANU' and not solicitud.analista_asignado_id and solicitud.usuario_creador_id:
            solicitud.analista_asignado = solicitud.usuario_creador

        if nuevo_estado in ('RECT', 'FIN'):
            solicitud.fecha_resolucion = timezone.now().date()

        solicitud.usuario_modificador = request.user
        solicitud.save()

        registrar_bitacora(
            solicitud=solicitud,
            usuario=request.user,
            estado_anterior=estado_anterior,
            estado_nuevo=nuevo_estado,
            accion=f"CAMBIO_ESTADO:{estado_anterior}->{nuevo_estado}",
            comentario=comentario,
            ip=get_client_ip(request),
        )

        # Envío de correo deshabilitado temporalmente
        # if nuevo_estado == 'PEND':
        #     from .email_utils import enviar_notificacion_solicitud
        #     enviar_notificacion_solicitud(solicitud)

        return Response(SolicitudDetailSerializer(solicitud, context={'request': request}).data)

    @action(detail=False, methods=['post'], url_path='reasignar')
    def reasignar(self, request):
        if request.user.rol not in ('ADMIN', 'SUPER'):
            return Response({'detail': 'Sin permiso para reasignar.'}, status=status.HTTP_403_FORBIDDEN)

        ids              = request.data.get('ids', [])
        nuevo_analista_id = request.data.get('analista_id')
        comentario       = request.data.get('comentario', '')

        if not ids or not nuevo_analista_id:
            return Response({'detail': 'Debe indicar ids y analista_id.'}, status=status.HTTP_400_BAD_REQUEST)

        from apps.accounts.models import CustomUser
        try:
            nuevo_analista = CustomUser.objects.get(pk=nuevo_analista_id, is_active=True)
        except CustomUser.DoesNotExist:
            return Response({'detail': 'Analista destino no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        qs    = Solicitud.objects.filter(pk__in=ids, estado__in=['ASIG', 'REV']).select_related('analista_asignado')
        ip    = get_client_ip(request)
        count = 0
        nuevo_nombre = nuevo_analista.get_full_name() or nuevo_analista.username

        for sol in qs:
            anterior_nombre = (
                sol.analista_asignado.get_full_name() or sol.analista_asignado.username
                if sol.analista_asignado else 'Sin asignar'
            )
            sol.analista_asignado   = nuevo_analista
            sol.asignado_por        = request.user
            sol.usuario_modificador = request.user
            sol.save()
            registrar_bitacora(
                solicitud=sol,
                usuario=request.user,
                estado_anterior=sol.estado,
                estado_nuevo=sol.estado,
                accion='REASIG',
                comentario=comentario or f'Reasignado de {anterior_nombre} a {nuevo_nombre}',
                ip=ip,
            )
            count += 1

        return Response({'detail': f'{count} solicitud(es) reasignada(s).', 'count': count})

    @action(detail=True, methods=['get'], url_path='bitacora')
    def bitacora(self, request, pk=None):
        solicitud = self.get_object()
        entradas  = solicitud.bitacora.select_related('usuario').order_by('-created_at')
        serializer = BitacoraSerializer(entradas, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'], url_path='transiciones')
    def transiciones(self, request, pk=None):
        solicitud   = self.get_object()
        disponibles = transiciones_disponibles(request.user.rol, solicitud.estado)
        return Response({'transiciones': disponibles})

    @action(detail=True, methods=['post'], url_path='enviar_notificacion')
    def enviar_notificacion(self, request, pk=None):
        solicitud = self.get_object()
        from .email_utils import enviar_notificacion_solicitud
        extras = request.data.get('destinatarios_extra', [])
        if isinstance(extras, str):
            extras = [e.strip() for e in extras.split(',') if e.strip()]
        ok, mensaje = enviar_notificacion_solicitud(solicitud, destinatarios_extra=extras or None)
        if ok:
            return Response({'detail': mensaje})
        return Response({'detail': mensaje}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='exportar')
    def exportar_excel(self, request):
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        qs = self.filter_queryset(self.get_queryset())
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Solicitudes'

        ORO_HEX  = 'CBab58'
        NAVY_HEX = '0F1932'

        header_fill   = PatternFill('solid', fgColor=ORO_HEX)
        header_font   = Font(bold=True, color=NAVY_HEX)
        header_align  = Alignment(horizontal='center', vertical='center')
        thin_border   = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        headers = [
            'N° Solicitud', 'Estado', 'Prioridad',
            'Asegurado', 'CI Asegurado', 'Empleador',
            'Tipo Causal', 'Regional', 'Analista Asignado',
            'Fecha Recepción', 'Fecha Límite', 'Monto Total',
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill    = header_fill
            cell.font    = header_font
            cell.alignment = header_align
            cell.border  = thin_border

        ws.row_dimensions[1].height = 22

        from .models import ESTADO_CHOICES, PRIORIDAD_CHOICES
        estado_map   = dict(ESTADO_CHOICES)
        prioridad_map = dict(PRIORIDAD_CHOICES)

        for row_idx, sol in enumerate(qs, start=2):
            row_data = [
                sol.numero_solicitud,
                estado_map.get(sol.estado, sol.estado),
                prioridad_map.get(sol.prioridad, sol.prioridad),
                sol.asegurado.nombre_completo if sol.asegurado else '',
                sol.asegurado.cedula if sol.asegurado else '',
                sol.empleador.nombre_razon_social if sol.empleador else '',
                sol.tipo_causal.nombre if sol.tipo_causal else '',
                sol.regional.nombre if sol.regional else '',
                sol.analista_asignado.get_full_name() if sol.analista_asignado else '',
                sol.fecha_recepcion,
                sol.fecha_limite,
                float(sol.monto_total),
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buffer   = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"solicitudes_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=['get'], url_path='formulario-regularizacion')
    def formulario_regularizacion(self, request, pk=None):
        import openpyxl, os, copy
        from decimal import Decimal

        sol = self.get_object()
        template_path = os.path.join(
            os.path.dirname(__file__),
            'templates_xlsx', 'formulario_regularizacion.xlsx'
        )
        wb = openpyxl.load_workbook(template_path)
        ld = wb['LLENAR DATOS']

        # ── Datos generales en LLENAR DATOS fila 3 ──────────────────────────
        ld['A3'] = sol.fecha_recepcion.strftime('%d/%m/%Y') if sol.fecha_recepcion else ''

        aseg = sol.asegurado
        if aseg:
            ld['B3'] = aseg.nombre_completo
            tipo_id_aseg = aseg.tipo_identificacion.codigo if aseg.tipo_identificacion else ''
            ld['C3'] = tipo_id_aseg
            ld['D3'] = aseg.cedula or ''
            ld['E3'] = aseg.cua or ''
        else:
            ld['B3'] = ld['C3'] = ld['D3'] = ld['E3'] = ''

        empl = sol.empleador
        if empl:
            ld['F3'] = empl.nombre_razon_social or ''
            tipo_id_empl = empl.tipo_identificacion.codigo if empl.tipo_identificacion else ''
            ld['G3'] = tipo_id_empl
            ld['H3'] = empl.numero_documento_identidad or ''
        else:
            ld['F3'] = ld['G3'] = ld['H3'] = ''

        ld['I3'] = sol.detalle_causal or ''

        # ── Periodos (formularios FPC) en FORM1 ─────────────────────────────
        fs = list(sol.formularios.all().order_by('periodo'))
        ws1 = wb['FORM1']

        # Mapa: 10 slots → (fila, col_offset) en FORM1
        # Cada slot ocupa 3 columnas: MES, AÑO, TOTAL_GANADO
        # Filas de datos exclusión: 29 (slots 0-4) y 31 (slots 5-9)
        excl_slots = [(29, 2), (29, 5), (29, 8), (29, 11), (29, 14),
                      (31, 2), (31, 5), (31, 8), (31, 11), (31, 14)]

        for i, form in enumerate(fs[:10]):
            row, col = excl_slots[i]
            periodo = str(form.periodo or '')
            # Soporta formatos: "MM/AAAA", "AAAA-MM", "MM-AAAA"
            if '/' in periodo:
                parts = periodo.split('/')
                mes, anio = (parts[0], parts[1]) if len(parts[0]) <= 2 else (parts[1], parts[0])
            elif '-' in periodo:
                parts = periodo.split('-')
                if len(parts[0]) == 4:
                    anio, mes = parts[0], parts[1]
                else:
                    mes, anio = parts[0], parts[1]
            else:
                mes, anio = periodo, ''
            ws1.cell(row=row, column=col,     value=mes)
            ws1.cell(row=row, column=col + 1, value=anio)
            ws1.cell(row=row, column=col + 2, value=float(form.total_ganado or 0))

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        num = sol.numero_solicitud
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="form_regularizacion_{num}.xlsx"'
        return response


class BitacoraViewSet(viewsets.ReadOnlyModelViewSet):
    queryset           = BitacoraSolicitud.objects.select_related('solicitud', 'usuario').order_by('-created_at')
    serializer_class   = BitacoraSerializer
    permission_classes = [IsAuthenticated]
    filter_backends    = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields   = ['solicitud', 'usuario', 'estado_nuevo']


class FormularioViewSet(SoftDeleteMixin, viewsets.ModelViewSet):
    queryset           = Formulario.objects.select_related('solicitud', 'tipo_planilla').all()
    serializer_class   = FormularioSerializer
    permission_classes = [IsAuthenticated]
    filter_backends    = [DjangoFilterBackend]
    filterset_fields   = ['solicitud', 'tipo_planilla']

    def perform_create(self, serializer):
        serializer.save(usuario_creador=self.request.user)


class AseguradoViewSet(SoftDeleteMixin, viewsets.ModelViewSet):
    queryset           = Asegurado.objects.all()
    serializer_class   = AseguradoSerializer
    permission_classes = [IsAuthenticated]
    filter_backends    = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields   = ['cedula', 'cua', 'tipo_identificacion']
    search_fields      = ['nombre', 'ap_paterno', 'ap_materno', 'cedula', 'cua']

    def perform_create(self, serializer):
        serializer.save(usuario_creador=self.request.user)


class EmpleadorViewSet(SoftDeleteMixin, viewsets.ModelViewSet):
    queryset           = Empleador.objects.all()
    serializer_class   = EmpleadorSerializer
    permission_classes = [IsAuthenticated]
    filter_backends    = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields   = ['numero_documento_identidad', 'nit', 'tipo_identificacion']
    search_fields      = ['nombre_razon_social', 'nit', 'numero_documento_identidad']

    def perform_create(self, serializer):
        serializer.save(usuario_creador=self.request.user)


def _seccion_doc(instance):
    obs = instance.observacion or ''
    if obs == '__NOTIF_ASE__':
        return 'Notificación Asegurado'
    if obs == '__NOTIF_EMP__':
        return 'Notificación Empleador'
    if obs == 'FORM_REGULARIZACION':
        return 'Formulario de Regularización'
    if instance.documento_id:
        nombre = instance.documento.descripcion if instance.documento else 'Documento de respaldo'
        return f'Documentación de Respaldo — {nombre}'
    return f'Form Interno — {obs}' if obs else 'Form Interno'


class DocumentoRespaldoViewSet(SoftDeleteMixin, viewsets.ModelViewSet):
    queryset           = DocumentoRespaldo.objects.select_related('documento', 'estado_documentacion', 'solicitud').all()
    serializer_class   = DocumentoRespaldoSerializer
    permission_classes = [IsAuthenticated]
    filter_backends    = [DjangoFilterBackend]
    filterset_fields   = ['solicitud', 'documento', 'estado_documentacion']
    parser_classes     = [parsers.MultiPartParser, parsers.FormParser, parsers.JSONParser]

    def perform_create(self, serializer):
        instance = serializer.save(usuario_creador=self.request.user)
        seccion  = _seccion_doc(instance)
        if instance.archivo:
            nombre_archivo = instance.archivo.name.split('/')[-1]
            accion     = 'DOC_SUBIDO'
            comentario = f'Archivo adjuntado en "{seccion}": {nombre_archivo}'
        elif instance.documento_id:
            doc_nombre = instance.documento.descripcion if instance.documento else 'Documento'
            accion     = 'DOC_RECIBIDO'
            comentario = f'Documento marcado como recibido: {doc_nombre}'
        else:
            return
        registrar_bitacora(
            solicitud=instance.solicitud,
            usuario=self.request.user,
            estado_anterior=instance.solicitud.estado,
            estado_nuevo=instance.solicitud.estado,
            accion=accion,
            comentario=comentario,
            ip=get_client_ip(self.request),
        )

    def perform_update(self, serializer):
        instance = serializer.save(usuario_modificador=self.request.user)
        if 'archivo' in serializer.validated_data and instance.archivo:
            seccion        = _seccion_doc(instance)
            nombre_archivo = instance.archivo.name.split('/')[-1]
            registrar_bitacora(
                solicitud=instance.solicitud,
                usuario=self.request.user,
                estado_anterior=instance.solicitud.estado,
                estado_nuevo=instance.solicitud.estado,
                accion='DOC_REEMPLAZADO',
                comentario=f'Archivo reemplazado en "{seccion}": {nombre_archivo}',
                ip=get_client_ip(self.request),
            )

    def destroy(self, request, *args, **kwargs):
        instance  = self.get_object()
        seccion   = _seccion_doc(instance)
        solicitud = instance.solicitud
        if instance.archivo:
            detalle = instance.archivo.name.split('/')[-1]
        elif instance.documento_id:
            detalle = instance.documento.descripcion if instance.documento else 'Documento'
        else:
            detalle = 'adjunto'
        instance.delete(user=request.user)
        registrar_bitacora(
            solicitud=solicitud,
            usuario=request.user,
            estado_anterior=solicitud.estado,
            estado_nuevo=solicitud.estado,
            accion='DOC_ELIMINADO',
            comentario=f'Documento eliminado de "{seccion}": {detalle}',
            ip=get_client_ip(request),
        )
        return Response(
            {'detail': 'Registro eliminado. Puede restaurarlo si fue un error.'},
            status=status.HTTP_200_OK,
        )


class SolicitanteViewSet(SoftDeleteMixin, viewsets.ModelViewSet):
    queryset           = Solicitante.objects.select_related('unidad', 'usuario').all()
    serializer_class   = SolicitanteSerializer
    permission_classes = [IsAuthenticated]
    filter_backends    = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields   = ['unidad']
    search_fields      = ['nombre', 'ap_paterno', 'ap_materno']

    def perform_create(self, serializer):
        serializer.save(usuario_creador=self.request.user)
